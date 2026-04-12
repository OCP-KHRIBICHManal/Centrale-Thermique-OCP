import streamlit as st
import pandas as pd
import openpyxl
from scipy.optimize import minimize
import plotly.graph_objects as go

def lire_excel_complet(fichier):
    wb = openpyxl.load_workbook(fichier, read_only=True)
    params = {}
    diagnostics = {}
    ws1 = wb['1_Parametres_Fixes']
    for row in ws1.iter_rows(values_only=True):
        if row[0] and isinstance(row[0], str) and isinstance(row[2], (int, float)):
            params[row[0].strip()] = float(row[2])
    ws6 = wb['6_Diagnostic_Solutions']
    for row in ws6.iter_rows(values_only=True):
        if row[0] and isinstance(row[0], str) and row[0][0] in ('Q','T','P'):
            diagnostics[row[0].strip()] = {
                'nominal' : float(row[1]) if isinstance(row[1], (int,float)) else None,
                'cause'   : str(row[4]) if row[4] else '—',
                'solution': str(row[5]) if row[5] else '—',
            }
    wb.close()
    return params, diagnostics

def calculer_puissance(x1, x4, x5, P):
    x6 = x1 - x4 - x5
    if x6 < 0:
        return 0.0, 0.0, x6
    h_HP    = P.get('h_HP', 3382.1)
    h_exh   = P.get('h_exhaust', 2285.0)
    h_MP_F1 = P.get('h_MP_bleed_F1', 3040.0)
    h_LP_F2 = P.get('h_LP_F2_bleed', 2886.6)
    eta     = P.get('eta_generateur', 0.985)
    P_meca  = (x1*(h_HP-h_exh) - x4*(h_MP_F1-h_exh) - x5*(h_LP_F2-h_exh)) / 3.6
    P_elec  = P_meca * eta
    return round(P_meca,1), round(P_elec,1), round(x6,2)

def bilan_complet(cadence, x1, x2, x3, x4, x5, D_MP, D_BP, Q_HRS, P):
    Q_HP_dispo = round((cadence/100)*P.get('Q_HP_max',217.0), 2)
    P_meca, P_elec, x6 = calculer_puissance(x1, x4, x5, P)
    h_HP    = P.get('h_HP', 3382.1)
    h_exh   = P.get('h_exhaust', 2285.0)
    h_MP_F1 = P.get('h_MP_bleed_F1', 3040.0)
    h_LP_F2 = P.get('h_LP_F2_bleed', 2886.6)
    ei = x1*(h_HP-148)/3.6
    eta_glob = round(P_elec/ei*100, 2) if ei > 0 else 0
    return {
        'Q_HP_dispo': Q_HP_dispo,
        'x1':x1,'x2':x2,'x3':x3,'x4':x4,'x5':x5,'x6':x6,
        'P_meca': P_meca, 'P_elec_kW': P_elec,
        'P_elec_MW': round(P_elec/1000,3),
        'eta_glob': eta_glob,
        'energie_entree': round(ei,1),
        'ecart_HP'  : round(Q_HP_dispo-(x1+x2+x3),3),
        'ecart_MP'  : round((x2+x4)-(D_MP+3+1.22),3),
        'ecart_BP'  : round((x3+x5+Q_HRS)-(D_BP+3),3),
        'ecart_turb': round(x1-(x4+x5+x6),3),
        'perte_bypass_MP' : round(x2*(h_HP-h_MP_F1)/3.6,1),
        'perte_bypass_BP' : round(x3*(h_HP-h_LP_F2)/3.6,1),
        'perte_condenseur': round(x6*(h_exh-148)/3.6,1),
        'perte_meca_elec' : round(P_meca-P_elec,1),
    }

def diagnostiquer(mesures, diag):
    res = []
    for param, val in mesures.items():
        info = diag.get(param, {})
        nom  = info.get('nominal')
        if not nom or nom == 0: continue
        try: ecart = (float(val)-float(nom))/abs(float(nom))*100
        except: continue
        statut = '✅ OK' if abs(ecart)<5 else ('⚠️ WARNING' if abs(ecart)<10 else '🔴 CRITIQUE')
        res.append({
            'Paramètre': param, 'Nominal': nom,
            'Mesuré': round(float(val),3), 'Écart (%)': round(ecart,2),
            'Statut': statut,
            'Cause'   : info.get('cause','—')    if abs(ecart)>=5 else '—',
            'Solution': info.get('solution','—') if abs(ecart)>=5 else '—',
        })
    return pd.DataFrame(res)

def optimiser(cadence, D_MP, D_BP, Q_HRS, P):
    Q_HP = (cadence/100)*P.get('Q_HP_max',217.0)
    def obj(x): return -calculer_puissance(x[0],x[3],x[4],P)[1]
    cstr = [
        {'type':'eq',   'fun': lambda x: x[0]+x[1]+x[2]-Q_HP},
        {'type':'ineq', 'fun': lambda x: (x[1]+x[3])-(D_MP+4.22)},
        {'type':'ineq', 'fun': lambda x: (x[2]+x[4]+Q_HRS)-(D_BP+3)},
        {'type':'ineq', 'fun': lambda x: x[0]-x[3]-x[4]},
        {'type':'ineq', 'fun': lambda x: calculer_puissance(x[0],x[3],x[4],P)[1]-P.get('P_elec_min',10000)},
        {'type':'ineq', 'fun': lambda x: P.get('P_elec_max',63000)-calculer_puissance(x[0],x[3],x[4],P)[1]},
    ]
    bornes = [(100,Q_HP),(0,Q_HP),(0,Q_HP),(0,25),(0,130)]
    x0 = [min(217,Q_HP),0,0,12,57.4]
    res = minimize(obj, x0, method='SLSQP', bounds=bornes, constraints=cstr,
                   options={'maxiter':1000,'ftol':1e-9})
    if res.success:
        x1,x2,x3,x4,x5 = [round(v,2) for v in res.x]
        return True,x1,x2,x3,x4,x5,round(-res.fun,1)
    return False,*x0,0.0

# ══════════════════════════════════════════════════════
st.set_page_config(page_title="EMS — Centrale OCP", page_icon="⚡", layout="wide")

st.markdown("""
<style>
.entete{background:#1F4E79;color:white;padding:1rem 1.5rem;border-radius:10px;margin-bottom:1rem;}
.entete h2{margin:0;font-size:1.4rem;}
.entete p{margin:4px 0 0;font-size:0.9rem;opacity:0.85;}
</style>
<div class='entete'>
<h2>⚡ EMS — Centrale Thermique OCP</h2>
<p>Simulation · Bilan énergétique · Diagnostic · Optimisation</p>
</div>
""", unsafe_allow_html=True)

P = {
    'h_HP':3382.1,'h_exhaust':2285.0,'h_MP_bleed_F1':3040.0,
    'h_LP_F2_bleed':2886.6,'eta_generateur':0.985,
    'Q_HP_max':217.0,'P_elec_min':10000.0,'P_elec_max':63000.0,
}
DIAG = {
    'Q_HP_entree'   :{'nominal':217,   'cause':'Fuite ligne HP ou vanne HV-007','solution':'Inspection tuyauterie HP + recalibrage HV-007'},
    'Q_turbine'     :{'nominal':219,   'cause':'Vanne HV-012 défaillante ou encrassée','solution':'Vérification + nettoyage vanne HV-012'},
    'Q_sout_MP'     :{'nominal':12,    'cause':'Vanne PV-004 soutirage MP mal réglée','solution':'Recalibrage ou remplacement vanne PV-004'},
    'Q_sout_LP'     :{'nominal':57.4,  'cause':'Vanne PV-024 soutirage LP mal réglée','solution':'Recalibrage ou remplacement vanne PV-024'},
    'Q_det_HP_MP'   :{'nominal':0.01,  'cause':'Bypass HP→MP ouvert → perte enthalpique directe','solution':'Fermer vanne PV-004 bypass'},
    'Q_det_HP_BP'   :{'nominal':0.01,  'cause':'Bypass HP→BP ouvert → perte enthalpique directe','solution':'Fermer vanne PV-024 bypass'},
    'P_elec'        :{'nominal':53446, 'cause':'Dégradation turbine ou génératrice','solution':'Inspection turbine GE — contrôle étanchéité'},
    'Q_refroid_cond':{'nominal':149.6, 'cause':'Encrassement faisceaux condenseur','solution':'Nettoyage faisceaux — curage chimique'},
    'P_cond'        :{'nominal':0.081, 'cause':'Perte de vide — fuite air ou pompes vide','solution':'Vérifier pompes à vide + étanchéité'},
    'T_cond'        :{'nominal':41.8,  'cause':'Eau refroidissement trop chaude','solution':'Vérifier circuit eau refroid. + tour'},
    'T_HP'          :{'nominal':470,   'cause':'Anomalie chaudière ou désurchauffe excessive','solution':'Contrôler désurchauffe TV-003/TV-023'},
    'P_HP'          :{'nominal':57,    'cause':'Chute pression HP — fuite amont','solution':'Inspection ligne HP + coordination SAP'},
}

with st.sidebar:
    st.header("📂 Importer ton Excel")
    fichier = st.file_uploader("Centrale_Thermique_Modele_COMPLET.xlsx", type=['xlsx'])
    if fichier:
        try:
            p_lus, d_lus = lire_excel_complet(fichier)
            P.update(p_lus)
            DIAG.update(d_lus)
            st.success(f"✅ {len(p_lus)} paramètres chargés depuis Excel")
        except Exception as e:
            st.error(f"Erreur : {e}")

    st.divider()
    st.subheader("⚙️ Paramètres")
    cadence = st.slider("Cadence SAP (%)", 0, 100, 100)
    Q_HP_d  = round((cadence/100)*P['Q_HP_max'], 1)
    st.info(f"Débit HP dispo : **{Q_HP_d} T/h**")
    x1 = st.number_input("x1 — Q_turbine (T/h)",   0.0, float(Q_HP_d), min(217.0,float(Q_HP_d)), 1.0)
    x2 = st.number_input("x2 — Bypass HP→MP (T/h)", 0.0, float(Q_HP_d), 0.0, 1.0)
    x3 = st.number_input("x3 — Bypass HP→BP (T/h)", 0.0, float(Q_HP_d), 0.0, 1.0)
    x4 = st.number_input("x4 — Soutirage MP (T/h)",  0.0, 25.0,  12.0, 1.0)
    x5 = st.number_input("x5 — Soutirage LP (T/h)",  0.0, 130.0, 57.4, 1.0)
    st.divider()
    st.subheader("👥 Clients")
    D_MP  = st.number_input("Demande MP (T/h)", 0.0, 50.0,  10.0, 1.0)
    D_BP  = st.number_input("Demande BP (T/h)", 0.0, 123.0, 80.0, 1.0)
    Q_HRS = st.number_input("Apport HRS (T/h)", 0.0, 72.0,  0.0,  1.0)

B = bilan_complet(cadence, x1, x2, x3, x4, x5, D_MP, D_BP, Q_HRS, P)

tab1, tab2, tab3, tab4, tab5 = st.tabs(["Bilan Énergétique", "Analyse des Pertes", "Diagnostic", "Optimisation", " Schéma P&ID"])
with tab1:
    st.header(" Bilan Énergétique")
    c1,c2,c3,c4 = st.columns(4)
    c1.metric(" Puissance",     f"{B['P_elec_MW']:.3f} MW", f"{B['P_elec_kW']:,.0f} kW")
    c2.metric(" Efficacité η",  f"{B['eta_glob']:.2f} %",  f"{B['eta_glob']-36:.2f}% vs optimal")
    c3.metric(" Condenseur x6", f"{B['x6']:.1f} T/h")
    c4.metric(" Énergie entrée",f"{B['energie_entree']:,.0f} kW")
    st.divider()
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Bilans massiques des 4 nœuds")
        df_b = pd.DataFrame({
            'Nœud'       : ['HP','MP','BP','Turbine'],
            'Équation'   : [
                f"Q_HP({B['Q_HP_dispo']}) = x1+x2+x3 ({x1+x2+x3:.1f})",
                f"x2+x4 ({x2+x4:.1f}) vs clients_MP ({D_MP+4.22:.2f})",
                f"x3+x5+HRS ({x3+x5+Q_HRS:.1f}) vs clients_BP ({D_BP+3:.1f})",
                f"x1({x1}) = x4+x5+x6 ({x4+x5+B['x6']:.1f})"
            ],
            'Écart (T/h)': [B['ecart_HP'],B['ecart_MP'],B['ecart_BP'],B['ecart_turb']]
        })
        def color_e(val):
            if not isinstance(val,float): return ''
            if abs(val)<0.1: return 'background-color:#d4edda'
            if abs(val)<2:   return 'background-color:#fff3cd'
            return 'background-color:#f8d7da'
        st.dataframe(df_b.style.map(color_e,subset=['Écart (T/h)']),
                     use_container_width=True, hide_index=True)
    with col2:
        st.subheader("Répartition débit vapeur HP")
        fig = go.Figure(go.Pie(
            labels=['Turbine x1','Bypass MP x2','Bypass BP x3'],
            values=[max(x1,0.01),max(x2,0.01),max(x3,0.01)],
            hole=0.4, marker_colors=['#1F4E79','#E67E22','#E74C3C']
        ))
        fig.update_layout(height=250,margin=dict(t=0,b=0,l=0,r=0))
        st.plotly_chart(fig, use_container_width=True)
    st.subheader("Flux énergétique (Sankey)")
    fig_s = go.Figure(go.Sankey(
        node=dict(label=["Vapeur HP","Turbine","Électricité","Condenseur","Soutirages","Bypass MP","Bypass BP"],
                  color=["#1F4E79","#27AE60","#F1C40F","#E74C3C","#8E44AD","#E67E22","#D35400"]),
        link=dict(source=[0,1,1,1,0,0],target=[1,2,3,4,5,6],
                  value=[max(x1,.1),max(B['P_elec_kW']/1000,.1),max(B['x6'],.1),
                         max(x4+x5,.1),max(x2,.1),max(x3,.1)])
    ))
    fig_s.update_layout(height=280,margin=dict(t=10,b=10,l=10,r=10))
    st.plotly_chart(fig_s, use_container_width=True)

with tab2:
    st.header("🔍 Pertes Enthalpiques")
    pertes = {'Condenseur':B['perte_condenseur'],'Bypass HP→MP':B['perte_bypass_MP'],
              'Bypass HP→BP':B['perte_bypass_BP'],'Méca→Élec':B['perte_meca_elec']}
    total = sum(pertes.values())
    c1,c2,c3 = st.columns(3)
    c1.metric("Énergie entrante",  f"{B['energie_entree']:,.0f} kW")
    c2.metric("Puissance utile",   f"{B['P_elec_kW']:,.0f} kW")
    c3.metric("Pertes totales",    f"{total:,.0f} kW", f"{total/max(B['energie_entree'],1)*100:.1f}%")
    col1,col2 = st.columns(2)
    with col1:
        fig_p = go.Figure(go.Bar(
            x=list(pertes.values()), y=list(pertes.keys()), orientation='h',
            marker_color=['#E74C3C','#E67E22','#D35400','#95A5A6'],
            text=[f"{v:,.0f} kW" for v in pertes.values()], textposition='outside'
        ))
        fig_p.update_layout(height=250,margin=dict(t=0,b=0,l=0,r=30))
        st.plotly_chart(fig_p, use_container_width=True)
    with col2:
        df_p = pd.DataFrame({
            'Source'    : list(pertes.keys()),
            'Perte (kW)': [round(v,1) for v in pertes.values()],
            'Perte (%)'  : [round(v/max(B['energie_entree'],1)*100,2) for v in pertes.values()]
        })
        st.dataframe(df_p, use_container_width=True, hide_index=True)
        perte_unit_MP = round((P.get('h_HP',3382.1)-P.get('h_MP_bleed_F1',3040.0))/3.6,1)
        perte_unit_BP = round((P.get('h_HP',3382.1)-P.get('h_LP_F2_bleed',2886.6))/3.6,1)
        st.info(f"Perte unitaire bypass HP→MP : **{perte_unit_MP} kW/T/h**\n\nPerte unitaire bypass HP→BP : **{perte_unit_BP} kW/T/h**\n\n→ Objectif : **x2 = 0 et x3 = 0**")

with tab3:
    st.header("🔧 Diagnostic — Causes & Solutions")
    st.info("Saisis les valeurs mesurées → diagnostic automatique avec causes et solutions de ton Excel.")
    col1,col2,col3 = st.columns(3)
    mes = {}
    with col1:
        st.markdown("**Débits (T/h)**")
        mes['Q_HP_entree']    = st.number_input("Q_HP_entree",    value=217.0,   step=1.0)
        mes['Q_turbine']      = st.number_input("Q_turbine",      value=219.0,   step=1.0)
        mes['Q_sout_MP']      = st.number_input("Q_sout_MP",      value=12.0,    step=0.5)
        mes['Q_sout_LP']      = st.number_input("Q_sout_LP",      value=57.4,    step=0.5)
    with col2:
        st.markdown("**Bypass & Puissance**")
        mes['Q_det_HP_MP']    = st.number_input("Q_det_HP_MP",    value=0.0,     step=0.5)
        mes['Q_det_HP_BP']    = st.number_input("Q_det_HP_BP",    value=0.0,     step=0.5)
        mes['P_elec']         = st.number_input("P_elec (kW)",    value=53446.0, step=100.0)
        mes['Q_refroid_cond'] = st.number_input("Q_refroid_cond", value=149.6,   step=1.0)
    with col3:
        st.markdown("**T° et Pressions**")
        mes['P_cond'] = st.number_input("P_cond (Bar)",value=0.081,step=0.001,format="%.3f")
        mes['T_cond'] = st.number_input("T_cond (°C)", value=41.8, step=0.5)
        mes['T_HP']   = st.number_input("T_HP (°C)",   value=470.0,step=1.0)
        mes['P_HP']   = st.number_input("P_HP (Bar)",  value=57.0, step=0.5)

    if st.button("🔍 Lancer le diagnostic", type="primary", use_container_width=True):
        df_d = diagnostiquer(mes, DIAG)
        if df_d.empty:
            st.warning("Aucune donnée à diagnostiquer.")
        else:
            nb_ok   = len(df_d[df_d['Statut']=='✅ OK'])
            nb_warn = len(df_d[df_d['Statut']=='⚠️ WARNING'])
            nb_crit = len(df_d[df_d['Statut']=='🔴 CRITIQUE'])
            c1,c2,c3 = st.columns(3)
            c1.metric("✅ Normal",  nb_ok)
            c2.metric("⚠️ Warning", nb_warn)
            c3.metric("🔴 Critique",nb_crit)
            fig_d = go.Figure(go.Bar(
                x=df_d['Paramètre'], y=df_d['Écart (%)'].abs(),
                marker_color=['#E74C3C' if '🔴' in s else '#E67E22' if '⚠️' in s else '#27AE60' for s in df_d['Statut']],
                text=df_d['Écart (%)'].round(1).astype(str)+'%', textposition='outside'
            ))
            fig_d.add_hline(y=5, line_dash='dash',line_color='orange',annotation_text='Warning 5%')
            fig_d.add_hline(y=10,line_dash='dash',line_color='red',  annotation_text='Critique 10%')
            fig_d.update_layout(height=280,margin=dict(t=30,b=10,l=10,r=10))
            st.plotly_chart(fig_d, use_container_width=True)
            def cr(row):
                if '🔴' in str(row['Statut']): return ['background-color:#ffe0e0']*len(row)
                if '⚠️' in str(row['Statut']): return ['background-color:#fff8dc']*len(row)
                return ['background-color:#e8f5e9']*len(row)
            st.dataframe(df_d.style.apply(cr,axis=1),use_container_width=True,hide_index=True)
            alertes = df_d[df_d['Statut']!='✅ OK']
            if not alertes.empty:
                st.subheader("⚠️ Alertes — Causes & Solutions")
                for _, row in alertes.iterrows():
                    with st.expander(f"{row['Statut']} — {row['Paramètre']} (écart : {row['Écart (%)']:+.1f}%)"):
                        ca,cb = st.columns(2)
                        ca.markdown(f"**Nominal :** `{row['Nominal']}`")
                        ca.markdown(f"**Mesuré :** `{row['Mesuré']}`")
                        cb.markdown(f"**🔎 Cause :** {row['Cause']}")
                        cb.markdown(f"**🔧 Solution :** {row['Solution']}")

with tab4:
    st.header("🚀 Optimisation")
    st.markdown("L'optimiseur cherche **x1..x5** qui **maximisent P_elec** en respectant toutes les contraintes.")
    col1,col2 = st.columns([1,2])
    with col1:
        cad_opt  = st.slider("Cadence (%)",50,100,int(cadence),key='co')
        D_MP_opt = st.number_input("Demande MP",0.0,50.0,D_MP,1.0,key='om')
        D_BP_opt = st.number_input("Demande BP",0.0,123.0,D_BP,1.0,key='ob')
        Q_HRS_opt= st.number_input("Apport HRS",0.0,72.0,Q_HRS,1.0,key='oh')
        if st.button("🚀 Optimiser",type="primary",use_container_width=True):
            with st.spinner("Calcul en cours..."):
                ok,ox1,ox2,ox3,ox4,ox5,P_opt = optimiser(cad_opt,D_MP_opt,D_BP_opt,Q_HRS_opt,P)
            if ok:
                st.session_state['opt']={'x1':ox1,'x2':ox2,'x3':ox3,'x4':ox4,'x5':ox5,'P_elec':P_opt,
                    'bilan':bilan_complet(cad_opt,ox1,ox2,ox3,ox4,ox5,D_MP_opt,D_BP_opt,Q_HRS_opt,P)}
                st.success(f"✅ P_elec = {P_opt/1000:.3f} MW")
            else:
                st.error("Optimisation non convergée.")
    with col2:
        if 'opt' in st.session_state:
            opt=st.session_state['opt']; ob=opt['bilan']
            df_cmp=pd.DataFrame({
                'Variable' :['x1 Q_turbine','x2 Bypass MP','x3 Bypass BP','x4 Sout.MP','x5 Sout.LP'],
                'Actuelle' :[x1,x2,x3,x4,x5],
                'Optimale' :[opt['x1'],opt['x2'],opt['x3'],opt['x4'],opt['x5']],
            })
            df_cmp['Différence']=df_cmp['Optimale']-df_cmp['Actuelle']
            st.dataframe(df_cmp,use_container_width=True,hide_index=True)
            c1,c2,c3=st.columns(3)
            c1.metric("P_elec optimale",f"{opt['P_elec']/1000:.3f} MW",f"+{(opt['P_elec']-B['P_elec_kW'])/1000:.3f} MW")
            c2.metric("η optimale",f"{ob['eta_glob']:.2f} %",f"+{ob['eta_glob']-B['eta_glob']:.2f} pts")
            c3.metric("x6 condenseur",f"{ob['x6']:.1f} T/h")
            fig_c=go.Figure()
            fig_c.add_trace(go.Bar(name='Actuelle',x=df_cmp['Variable'],y=df_cmp['Actuelle'],marker_color='#95A5A6'))
            fig_c.add_trace(go.Bar(name='Optimale',x=df_cmp['Variable'],y=df_cmp['Optimale'],marker_color='#27AE60'))
            fig_c.update_layout(barmode='group',height=270,margin=dict(t=10,b=10,l=10,r=10))
            st.plotly_chart(fig_c,use_container_width=True)
        else:
            st.info("⬅️ Lance l'optimisation pour voir les résultats.")
with tab5:
    st.subheader("🗺️ Schéma P&ID — Réseau Vapeur HP/MP/BP")
    st.caption("Cliquez sur un tag PI pour voir ses détails")
    try:
        with open("schema_centrale_tags.html", "r", encoding="utf-8") as f:
            html_content = f.read()
        st.components.v1.html(html_content, height=950, scrolling=True)
    except:
        st.error("Fichier schema_centrale_tags.html introuvable")