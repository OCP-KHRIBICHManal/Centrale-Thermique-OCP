import streamlit as st
import pandas as pd
import numpy as np
from scipy.optimize import minimize
import plotly.graph_objects as go
import plotly.express as px
from openpyxl import load_workbook

# ─────────────────────────────────────────────
#  CONFIGURATION PAGE
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="EMS — Centrale Thermique OCP",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────
#  CSS PERSONNALISÉ
# ─────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #0d3a5c, #1a5276);
        padding: 1rem 1.5rem;
        border-radius: 10px;
        margin-bottom: 1.5rem;
        border-left: 4px solid #4fc3f7;
    }
    .main-header h1 { color: #4fc3f7; font-size: 1.4rem; margin: 0; }
    .main-header p  { color: #aed6f1; font-size: 0.85rem; margin: 0.3rem 0 0 0; }

    .metric-card {
        background: #f8f9fa;
        border: 1px solid #dee2e6;
        border-radius: 8px;
        padding: 1rem;
        text-align: center;
    }
    .kpi-ok      { border-left: 4px solid #28a745; }
    .kpi-warn    { border-left: 4px solid #ffc107; }
    .kpi-danger  { border-left: 4px solid #dc3545; }

    .tag-label {
        font-family: monospace;
        font-size: 0.78rem;
        background: #e8f4fd;
        color: #0d3a5c;
        padding: 2px 6px;
        border-radius: 4px;
        border: 1px solid #b8d4e8;
    }
    .section-title {
        font-size: 1rem;
        font-weight: 600;
        color: #0d3a5c;
        border-bottom: 2px solid #4fc3f7;
        padding-bottom: 0.3rem;
        margin-bottom: 1rem;
    }
    div[data-testid="stExpander"] { border: 1px solid #dee2e6; border-radius: 8px; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
#  PARAMÈTRES FIXES (GE Nuovo Pignone RATED)
# ─────────────────────────────────────────────
P = {
    # HP
    "P_HP":            57.0,      # Bar
    "T_HP":           470.0,      # °C
    "h_HP":          3382.1,      # kJ/kg
    "Q_HP_max":       217.0,      # T/h
    # Turbine
    "eta_gen":          0.985,
    "P_elec_min":   10000.0,      # kW
    "P_elec_max":   63000.0,      # kW
    "W_turbine":     3000.0,      # RPM
    # Soutirage MP (Flange 1)
    "P_collect_MP":    12.0,      # Bar
    "T_collect_MP":   270.0,      # °C
    "h_MP":          2978.0,      # kJ/kg
    "h_MP_F1":       3040.0,      # kJ/kg
    "Q_MP_bleed_nom":  12.0,      # T/h
    "Q_min_sout_MP":  200.0,      # T/h  débit min turbine si MP actif
    # Soutirage LP (Flanges 2+3)
    "P_collect_BP":     5.0,      # Bar
    "T_collect_BP":   190.0,      # °C
    "h_BP":          2855.0,      # kJ/kg
    "h_LP_F2":       2886.6,      # kJ/kg
    "Q_LP_bleed_nom":  57.4,      # T/h
    "Q_min_sout_LP":  100.0,      # T/h  débit min turbine si LP actif
    # Condenseur
    "P_cond":          0.081,     # Bar a
    "T_cond":          41.8,      # °C
    "h_exhaust":     2285.0,      # kJ/kg
    # Clients fixes
    "Q_DAP_MP":         3.0,      # T/h
    "Q_DAP_BP":         3.0,      # T/h
    "Q_ejecteur":       1.22,     # T/h
}

# ─────────────────────────────────────────────
#  DONNÉES DIAGNOSTIC (43 tags PI)
# ─────────────────────────────────────────────
TAGS = {
    # Format : code_python : {nom, tag_PI, nominal, unite, cause, solution}
    "P_HP_mes":               {"nom":"Pression vapeur HP",          "tag":"515APG10.PIC-328","nominal":57,    "unite":"Bar",    "cat":"HP"},
    "T_HP_mes":               {"nom":"Température vapeur HP",       "tag":"515APG10.TIC-327","nominal":470,   "unite":"°C",     "cat":"HP"},
    "Q_HP_mes":               {"nom":"Débit HP entrant (SAP)",      "tag":"515APG10.FI-203", "nominal":217,   "unite":"T/h",    "cat":"HP"},
    "P_elec_mes":             {"nom":"Puissance électrique",        "tag":"515APG10.JT-963A","nominal":53446, "unite":"kW",     "cat":"TRB"},
    "Q_HP_turb_mes":          {"nom":"Débit HP entrée turbine",     "tag":"515APG10.FI-151", "nominal":None,  "unite":"T/h",    "cat":"TRB"},
    "T_HP_turb_mes":          {"nom":"T° HP entrée turbine",        "tag":"515APG10.TI-154", "nominal":None,  "unite":"°C",     "cat":"TRB"},
    "P_HP_turb_mes":          {"nom":"P vapeur HP entrée turbine",  "tag":"515APG10.PI-153", "nominal":57,    "unite":"Bar",    "cat":"TRB"},
    "N_turb_mes":             {"nom":"Vitesse rotation turbine",    "tag":"515APG10.SE-288", "nominal":3000,  "unite":"RPM",    "cat":"TRB"},
    "Q_sout_MP_mes":          {"nom":"Débit soutirage MP",          "tag":"515APG10.FI-541", "nominal":12,    "unite":"T/h",    "cat":"MP"},
    "Q_sout_LP_mes":          {"nom":"Débit soutirage LP",          "tag":"515APG10.FI-104", "nominal":57.4,  "unite":"T/h",    "cat":"BP"},
    "Q_det_MP_mes":           {"nom":"Débit bypass HP→MP",          "tag":"515APG10.FI-005", "nominal":0,     "unite":"T/h",    "cat":"BP_MP"},
    "P_det_MP_mes":           {"nom":"Pression détente bypass MP",  "tag":"515APG10.PIC-004","nominal":12,    "unite":"Bar",    "cat":"BP_MP"},
    "T_det_MP_mes":           {"nom":"T° détente bypass MP",        "tag":"515APG10.TIC-003","nominal":270,   "unite":"°C",     "cat":"BP_MP"},
    "P_det_BP_mes":           {"nom":"Pression détente HP→BP",      "tag":"515APG10.PI-032", "nominal":5,     "unite":"Bar",    "cat":"BP_BP"},
    "T_det_BP_mes":           {"nom":"T° détente bypass BP",        "tag":"515APG10.TIC-023","nominal":190,   "unite":"°C",     "cat":"BP_BP"},
    "P_det_MP_BP_mes":        {"nom":"Pression détente MP→BP",      "tag":"515APG10.PIC-044","nominal":5,     "unite":"Bar",    "cat":"BP_BP"},
    "Q_HRS_mes":              {"nom":"Débit BP entrant HRS",        "tag":"515APG10.FI370",  "nominal":None,  "unite":"T/h",    "cat":"BP_BP"},
    "P_cond_mes":             {"nom":"Pression condenseur",         "tag":"515APG10.PI-252", "nominal":0.081, "unite":"Bar a",  "cat":"COND"},
    "T_cond_mes":             {"nom":"Température condenseur",      "tag":"515APG10.TI-170", "nominal":41.8,  "unite":"°C",     "cat":"COND"},
    "pct_PV004_mes":          {"nom":"Ouverture PV-004 (bypass MP)","tag":"515APG10.PV-004", "nominal":0,     "unite":"%",      "cat":"VANNE"},
    "pct_PV548_mes":          {"nom":"Ouverture PV-548",            "tag":"515APG10.PV-548", "nominal":None,  "unite":"%",      "cat":"VANNE"},
    "pct_PV024_mes":          {"nom":"Ouverture PV-024 (bypass BP)","tag":"515APG10.PV-024", "nominal":0,     "unite":"%",      "cat":"VANNE"},
    "pct_PV551B_mes":         {"nom":"Ouverture PV-551B",           "tag":"515APG10.PV-551B","nominal":None,  "unite":"%",      "cat":"VANNE"},
    "pct_PV044_mes":          {"nom":"Ouverture PV044 (MP→BP)",     "tag":"515APG10.PV044",  "nominal":0,     "unite":"%",      "cat":"VANNE"},
    "pct_TV003_mes":          {"nom":"Arrosage TV-003 bypass MP",   "tag":"515APG10.TV-003", "nominal":None,  "unite":"%",      "cat":"ARROS"},
    "pct_TV083_mes":          {"nom":"Arrosage TV-083 soutirée MP", "tag":"515APG10.TV-083", "nominal":None,  "unite":"%",      "cat":"ARROS"},
    "pct_TV023_mes":          {"nom":"Arrosage TV-023 bypass BP",   "tag":"515APG10.TV-023", "nominal":None,  "unite":"%",      "cat":"ARROS"},
    "Q_arros_bypass_MP_mes":  {"nom":"Débit arrosage bypass MP",    "tag":"515APG10.FI-001", "nominal":None,  "unite":"m³/h",   "cat":"ARROS"},
    "Q_arros_bypass_BP_mes":  {"nom":"Débit arrosage bypass BP",    "tag":"515APG10.FI021",  "nominal":None,  "unite":"m³/h",   "cat":"ARROS"},
    "Q_arros_MP_sout_mes":    {"nom":"Débit arrosage MP soutirée",  "tag":"515APG10.FI-081", "nominal":None,  "unite":"m³/h",   "cat":"ARROS"},
    "Q_DAP_MP_mes":           {"nom":"Débit DAP MP",                "tag":"515APG10.FI-066", "nominal":3,     "unite":"T/h",    "cat":"CONSO"},
    "P_DAP_MP_mes":           {"nom":"Pression DAP MP",             "tag":"515APG10.PI-065", "nominal":12,    "unite":"Bar",    "cat":"CONSO"},
    "T_DAP_MP_mes":           {"nom":"Température DAP MP",          "tag":"515APG10.TIC-063","nominal":270,   "unite":"°C",     "cat":"CONSO"},
    "Q_JPH_recep_mes":        {"nom":"Débit réception JPH MP",      "tag":"515APG10.FI-043A","nominal":0,     "unite":"T/h",    "cat":"CONSO"},
    "Q_JPH_transf_mes":       {"nom":"Débit transfert JPH MP",      "tag":"515APG10.FI-043B","nominal":0,     "unite":"T/h",    "cat":"CONSO"},
    "P_JPH_MP_mes":           {"nom":"Pression JPH MP",             "tag":"515APG10.PI-041", "nominal":12,    "unite":"Bar",    "cat":"CONSO"},
    "T_JPH_MP_mes":           {"nom":"Température JPH MP",          "tag":"515APG10.TI-042", "nominal":270,   "unite":"°C",     "cat":"CONSO"},
    "Q_CAP_BP_mes":           {"nom":"Débit vapeur BP CAP",         "tag":"515APG10.FI-104", "nominal":0,     "unite":"T/h",    "cat":"CONSO"},
    "P_CAP_BP_mes":           {"nom":"Pression vapeur BP CAP",      "tag":"515APG10.PI-106", "nominal":5,     "unite":"Bar",    "cat":"CONSO"},
    "T_CAP_BP_mes":           {"nom":"Température vapeur BP CAP",   "tag":"515APG10.TIC-103","nominal":190,   "unite":"°C",     "cat":"CONSO"},
    "pct_ATM1_mes":           {"nom":"Vanne sécurité ATM 1",        "tag":"515APG10.PIC-128A","nominal":0,    "unite":"%",      "cat":"ATM"},
    "pct_ATM2_mes":           {"nom":"Vanne sécurité ATM 2",        "tag":"515APG10.PIC-128B","nominal":0,    "unite":"%",      "cat":"ATM"},
}

DIAG = {
    "P_HP_mes":       {"cause":"Chute de pression HP amont — fuite ligne principale ou problème SAP",           "sol":"1) Inspection tuyauterie HP  2) Check vanne HV-007  3) Coordination unité sulfurique  4) Recalibrer PIC-328"},
    "T_HP_mes":       {"cause":"Désurchauffe excessive (TV-003/TV-023) ou anomalie chaudière amont",            "sol":"1) Contrôler désurchauffeur TV-003/TV-023  2) Vérifier chaudière amont  3) Inspecter isolation  4) Recalibrer TIC-327"},
    "Q_HP_mes":       {"cause":"Cadence SAP réduite ou fuite collecteur HP",                                    "sol":"1) Vérifier cadence SAP (objectif 100%)  2) Inspection collecteur HP  3) Recalibrer FI-203"},
    "P_elec_mes":     {"cause":"Dégradation interne turbine ou rendement génératrice réduit",                   "sol":"1) Inspection turbine GE  2) Mesurer η génératrice  3) Réduire bypasses  4) Planifier maintenance"},
    "Q_HP_turb_mes":  {"cause":"Vanne HV-012 défaillante ou débitmètre FI-151 faussé",                         "sol":"1) Vérification vanne HV-012  2) Recalibrer FI-151  3) Comparer avec FI-203"},
    "T_HP_turb_mes":  {"cause":"Perte thermique avant turbine ou thermocouple TI-154 défaillant",               "sol":"1) Vérifier isolation ligne HP  2) Recalibrer TI-154  3) Comparer TIC-327 vs TI-154"},
    "P_HP_turb_mes":  {"cause":"Chute pression localisée avant turbine — filtre colmaté ou capteur mal étalonné","sol":"1) Inspecter filtre admission turbine  2) Recalibrer PI-153  3) Comparer PIC-328 vs PI-153"},
    "N_turb_mes":     {"cause":"Régulateur vitesse GE défaillant ou surcharge réseau",                          "sol":"1) Vérifier régulateur vitesse GE (setpoint 3000 RPM)  2) Contrôler charge réseau  3) Recalibrer SE-288"},
    "Q_sout_MP_mes":  {"cause":"Vanne PV-004 mal réglée ou demande client MP anormale",                         "sol":"1) Recalibrer vanne PV-004  2) Vérifier demande clients MP  3) Recalibrer FI-541"},
    "Q_sout_LP_mes":  {"cause":"Vanne PV-024 mal réglée ou demande clients BP anormale",                        "sol":"1) Recalibrer vanne PV-024  2) Vérifier demande clients BP (max 123 T/h)  3) Recalibrer FI-104"},
    "Q_det_MP_mes":   {"cause":"Bypass HP→MP ouvert inutilement → PERTE ÉNERGÉTIQUE DIRECTE",                  "sol":"1) Vérifier si clients MP satisfaits par soutirage  2) Fermer PV-004 si possible  3) Inspecter fuite interne"},
    "P_det_MP_mes":   {"cause":"Régulateur PIC-004 mal étalonné — pression non régulée à 12 Bar",              "sol":"1) Recalibrer PIC-004  2) Vérifier consigne (cible 12 Bar)  3) Inspecter vanne PV-004"},
    "T_det_MP_mes":   {"cause":"Désurchauffe TV-003 insuffisante → T° sortie trop haute",                       "sol":"1) Augmenter ouverture TV-003  2) Cible T° = 270°C  3) Recalibrer TIC-003  4) Vérifier FI-001"},
    "P_det_BP_mes":   {"cause":"Capteur PI-032 mal étalonné ou détente n'atteint pas 5 Bar",                   "sol":"1) Recalibrer PI-032  2) Vérifier consigne bypass BP (cible 5 Bar)  3) Inspecter PV-024"},
    "T_det_BP_mes":   {"cause":"Désurchauffe TV-023 insuffisante → T° trop haute pour réseau BP",              "sol":"1) Augmenter ouverture TV-023  2) Cible T° = 190°C  3) Recalibrer TIC-023  4) Vérifier FI021"},
    "P_cond_mes":     {"cause":"Perte de vide condenseur : fuite air, pompes vide défaillantes",                "sol":"1) Test étanchéité condenseur  2) Contrôler pompes à vide  3) Purger gaz incondensables  4) Recalibrer PI-252"},
    "T_cond_mes":     {"cause":"Eau refroidissement trop chaude ou encrassement faisceaux",                     "sol":"1) Vérifier circuit eau refroidissement  2) Curage chimique faisceaux  3) Recalibrer TI-170"},
    "pct_PV004_mes":  {"cause":"Bypass MP ouvert inutilement → vapeur HP détendue sans travail = PERTE",       "sol":"1) Fermer PV-004 si clients MP satisfaits  2) Objectif = 0%  3) Vérifier fuite interne"},
    "pct_PV024_mes":  {"cause":"Bypass BP ouvert inutilement → vapeur HP détendue sans travail = PERTE",       "sol":"1) Fermer PV-024 si clients BP satisfaits  2) Objectif = 0%  3) Vérifier fuite interne"},
    "pct_ATM1_mes":   {"cause":"🚨 CRITIQUE : Surpression réseau vapeur ou fuite vers atmosphère",              "sol":"⚠️ URGENT : 1) Identifier source surpression  2) Intervention maintenance  3) Vérifier P réseau HP/MP/BP"},
    "pct_ATM2_mes":   {"cause":"🚨 CRITIQUE : Surpression réseau vapeur ou fuite vers atmosphère",              "sol":"⚠️ URGENT : 1) Identifier source surpression  2) Intervention maintenance  3) Vérifier P réseau HP/MP/BP"},
    "Q_DAP_MP_mes":   {"cause":"Consommation DAP MP anormale ou vanne alimentation mal réglée",                 "sol":"1) Vérifier état DAP MP  2) Recalibrer vanne alimentation  3) Recalibrer FI-066"},
    "P_DAP_MP_mes":   {"cause":"Pression MP au DAP anormale — chute réseau ou régulateur défaillant",          "sol":"1) Vérifier pression collecteur MP (12 Bar)  2) Inspecter régulateur DAP  3) Recalibrer PI-065"},
    "T_DAP_MP_mes":   {"cause":"Température MP au DAP hors gamme",                                              "sol":"1) Vérifier T° collecteur MP (cible 270°C)  2) Ajuster désurchauffe  3) Recalibrer TIC-063"},
    "P_cond_mes":     {"cause":"Perte de vide condenseur",                                                      "sol":"1) Test étanchéité condenseur  2) Contrôler pompes à vide  3) Purger gaz incondensables"},
}

# ─────────────────────────────────────────────
#  FONCTIONS THERMODYNAMIQUES
# ─────────────────────────────────────────────
def calc_puissance(x1, x4, x5):
    """Puissance électrique aux bornes (kW) — modèle simplifié GE"""
    if x1 <= 0:
        return 0.0
    # Bilan enthalpique turbine
    W_meca = (
        x1 * (P["h_HP"] - P["h_exhaust"])
        - x4 * (P["h_MP_F1"] - P["h_exhaust"])
        - x5 * (P["h_LP_F2"] - P["h_exhaust"])
    ) / 3.6  # conversion T/h × kJ/kg → kW
    P_elec = W_meca * P["eta_gen"]
    return max(0.0, P_elec)

def calc_energie_entree(x1, x2, x3):
    """Énergie enthalpique entrante vapeur HP (kW)"""
    Q_total = x1 + x2 + x3
    return Q_total * P["h_HP"] / 3.6

def calc_efficacite(x1, x2, x3, x4, x5):
    """Efficacité énergétique globale (%)"""
    E_entree = calc_energie_entree(x1, x2, x3)
    if E_entree <= 0:
        return 0.0
    P_elec = calc_puissance(x1, x4, x5)
    return (P_elec / E_entree) * 100.0

def bilan_complet(cadence, x1, x2, x3, x4, x5, D_MP, D_BP, Q_HRS):
    """Calcule tous les bilans et KPIs"""
    Q_HP_dispo = round((cadence / 100.0) * P["Q_HP_max"], 2)
    x6 = max(0.0, x1 - x4 - x5)

    P_elec    = calc_puissance(x1, x4, x5)
    E_entree  = calc_energie_entree(x1, x2, x3)
    eta_glob  = calc_efficacite(x1, x2, x3, x4, x5)

    bilan_HP    = x1 + x2 + x3 - Q_HP_dispo
    bilan_turb  = x1 - x4 - x5 - x6
    alim_MP     = x2 + x4
    alim_BP     = x3 + x5 + Q_HRS
    ecart_MP    = alim_MP - D_MP
    ecart_BP    = alim_BP - D_BP

    return {
        "Q_HP_dispo":  Q_HP_dispo,
        "x6":          x6,
        "P_elec_kW":   P_elec,
        "P_elec_MW":   P_elec / 1000.0,
        "E_entree_kW": E_entree,
        "eta_glob":    eta_glob,
        "bilan_HP":    bilan_HP,
        "bilan_turb":  bilan_turb,
        "alim_MP":     alim_MP,
        "alim_BP":     alim_BP,
        "ecart_MP":    ecart_MP,
        "ecart_BP":    ecart_BP,
    }

# ─────────────────────────────────────────────
#  OPTIMISEUR SCIPY
# ─────────────────────────────────────────────
def optimiser(cadence, D_MP, D_BP, Q_HRS):
    """Maximise l'efficacité énergétique sous contraintes"""
    Q_HP_dispo = (cadence / 100.0) * P["Q_HP_max"]

    def objectif(x):
        x1, x2, x3, x4, x5 = x
        return -calc_efficacite(x1, x2, x3, x4, x5)

    contraintes = [
        {"type": "eq",   "fun": lambda x: x[0] + x[1] + x[2] - Q_HP_dispo},
        {"type": "ineq", "fun": lambda x: x[1] + x[3] - D_MP},
        {"type": "ineq", "fun": lambda x: x[2] + x[4] + Q_HRS - D_BP},
        {"type": "ineq", "fun": lambda x: x[0] - x[3] - x[4]},
    ]

    bornes = [
        (100, Q_HP_dispo),
        (0, Q_HP_dispo),
        (0, Q_HP_dispo),
        (0, 25),
        (0, 130),
    ]

    x0 = [min(217, Q_HP_dispo * 0.9), 0, 0, 12, 57.4]

    try:
        res = minimize(objectif, x0, method="SLSQP",
                       bounds=bornes, constraints=contraintes,
                       options={"ftol": 1e-8, "maxiter": 1000})
        if res.success:
            x1, x2, x3, x4, x5 = res.x
            return {
                "success":  True,
                "x1": round(x1, 2), "x2": round(x2, 2), "x3": round(x3, 2),
                "x4": round(x4, 2), "x5": round(x5, 2),
                "x6": round(max(0, x1 - x4 - x5), 2),
                "P_elec_MW": round(calc_puissance(x1, x4, x5) / 1000, 3),
                "eta_opt":   round(calc_efficacite(x1, x2, x3, x4, x5), 2),
            }
    except Exception as e:
        pass
    return {"success": False}

# ─────────────────────────────────────────────
#  DIAGNOSTIC AUTOMATIQUE
# ─────────────────────────────────────────────
def diagnostiquer(mesures):
    """Compare mesures vs nominaux — retourne liste d'alertes"""
    alertes = []
    for code, val in mesures.items():
        if val is None or code not in TAGS:
            continue
        info = TAGS[code]
        nominal = info["nominal"]
        if nominal is None or nominal == 0:
            continue
        ecart = (val - nominal) / abs(nominal) * 100.0
        if abs(ecart) >= 10:
            statut = "🔴 CRITIQUE"
            cls = "danger"
        elif abs(ecart) >= 5:
            statut = "⚠️ WARNING"
            cls = "warn"
        else:
            statut = "✅ OK"
            cls = "ok"
        cause = DIAG.get(code, {}).get("cause", "—")
        sol   = DIAG.get(code, {}).get("sol",   "—")
        alertes.append({
            "code":    code,
            "nom":     info["nom"],
            "tag":     info["tag"],
            "nominal": nominal,
            "mesure":  val,
            "ecart":   round(ecart, 2),
            "statut":  statut,
            "classe":  cls,
            "unite":   info["unite"],
            "cause":   cause,
            "sol":     sol,
        })
    return sorted(alertes, key=lambda x: abs(x["ecart"]), reverse=True)

# ─────────────────────────────────────────────
#  EN-TÊTE
# ─────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <h1>⚡ EMS — Système de Gestion Énergétique | Centrale Thermique OCP</h1>
  <p>Turbine GE Nuovo Pignone 3000 RPM — Réseau vapeur HP/MP/BP — UNITE 515A</p>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.header("📁 Données Excel")
    fichier = st.file_uploader("Importer Centrale_Thermique_Modele_COMPLET.xlsx",
                                type=["xlsx"])
    if fichier:
        try:
            wb = load_workbook(fichier, read_only=True)
            st.success(f"✅ Fichier chargé — {len(wb.sheetnames)} feuilles")
        except Exception as e:
            st.error(f"Erreur : {e}")

    st.divider()
    st.header("⚙️ Paramètres simulation")

    cadence = st.slider("Cadence SAP (%)", 0, 100, 100)
    Q_HP_dispo = round((cadence / 100.0) * P["Q_HP_max"], 1)
    st.info(f"Débit HP disponible : **{Q_HP_dispo} T/h**")

    st.subheader("Demandes clients")
    D_MP = st.number_input("Demande MP totale (T/h)", 0.0, 100.0, 16.22, 0.1)
    D_BP = st.number_input("Demande BP totale (T/h)", 0.0, 200.0, 123.0, 1.0)
    Q_HRS = st.number_input("Apport HRS vers BP (T/h)", 0.0, 72.0, 0.0, 0.5)

    st.divider()
    st.subheader("Variables de décision")
    x1 = st.number_input("x1 — Débit turbine (T/h)", 0.0, float(Q_HP_dispo), min(217.0, Q_HP_dispo), 1.0)
    x2 = st.number_input("x2 — Bypass HP→MP (T/h)", 0.0, float(Q_HP_dispo), 0.0, 0.5)
    x3 = st.number_input("x3 — Bypass HP→BP (T/h)", 0.0, float(Q_HP_dispo), 0.0, 0.5)
    x4 = st.number_input("x4 — Soutirage MP (T/h)", 0.0, 25.0, 12.0, 0.5)
    x5 = st.number_input("x5 — Soutirage LP (T/h)", 0.0, 130.0, 57.4, 0.5)

# ─────────────────────────────────────────────
#  CALCUL BILAN
# ─────────────────────────────────────────────
B = bilan_complet(cadence, x1, x2, x3, x4, x5, D_MP, D_BP, Q_HRS)
P_elec_opt_ref = calc_puissance(Q_HP_dispo, 12, 57.4)

# ─────────────────────────────────────────────
#  ONGLETS
# ─────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    " Bilan Énergétique",
    " Analyse des Pertes",
    " Diagnostic Tags PI",
    " Optimisation",
    " Schéma P&ID"
])

# ══════════════════════════════════════════════
# TAB 1 — BILAN ÉNERGÉTIQUE
# ══════════════════════════════════════════════
with tab1:
    st.markdown('<p class="section-title">Bilan Énergétique — Résultats simulation</p>', unsafe_allow_html=True)

    # KPIs principaux
    c1, c2, c3, c4 = st.columns(4)
    c1.metric(" Puissance produite",  f"{B['P_elec_MW']:.3f} MW",
              f"{B['P_elec_kW']:,.0f} kW")
    c2.metric(" Efficacité globale η", f"{B['eta_glob']:.2f} %",
              f"{B['eta_glob'] - 27.4:+.2f}% vs nominal")
    c3.metric(" Vers condenseur x6",  f"{B['x6']:.1f} T/h")
    c4.metric(" Énergie entrante HP",  f"{B['E_entree_kW']/1000:.1f} MW")

    st.divider()

    col1, col2 = st.columns(2)

    with col1:
        st.markdown('<p class="section-title">Bilans massiques — 4 nœuds</p>', unsafe_allow_html=True)

        df_bilan = pd.DataFrame([
            {"Nœud": "HP", "Entrée (T/h)": B["Q_HP_dispo"],
             "Sortie (T/h)": round(x1+x2+x3, 2),
             "Écart (T/h)": round(B["bilan_HP"], 3)},
            {"Nœud": "Turbine", "Entrée (T/h)": round(x1, 2),
             "Sortie (T/h)": round(x4+x5+B["x6"], 2),
             "Écart (T/h)": round(B["bilan_turb"], 3)},
            {"Nœud": "MP (12 Bar)", "Entrée (T/h)": round(B["alim_MP"], 2),
             "Sortie (T/h)": round(D_MP, 2),
             "Écart (T/h)": round(B["ecart_MP"], 2)},
            {"Nœud": "BP (5 Bar)", "Entrée (T/h)": round(B["alim_BP"], 2),
             "Sortie (T/h)": round(D_BP, 2),
             "Écart (T/h)": round(B["ecart_BP"], 2)},
        ])

        def color_ecart(val):
            if not isinstance(val, float): return ''
            if abs(val) < 0.1: return 'background-color:#d4edda'
            if abs(val) < 2:   return 'background-color:#fff3cd'
            return 'background-color:#f8d7da'

        st.dataframe(
            df_bilan.style.map(color_ecart, subset=["Écart (T/h)"]),
            use_container_width=True, hide_index=True
        )

    with col2:
        st.markdown('<p class="section-title">Répartition débit vapeur HP</p>', unsafe_allow_html=True)
        fig_pie = go.Figure(go.Pie(
            labels=["Turbine x1", "Bypass MP x2", "Bypass BP x3"],
            values=[max(x1, 0.01), max(x2, 0.01), max(x3, 0.01)],
            hole=0.4,
            marker_colors=["#1F4E79", "#E8903A", "#C0392B"]
        ))
        fig_pie.update_layout(
            showlegend=True, height=280,
            margin=dict(l=10, r=10, t=10, b=10),
            legend=dict(font=dict(size=11))
        )
        st.plotly_chart(fig_pie, use_container_width=True)

    # Tableau variables de décision
    st.divider()
    st.markdown('<p class="section-title">Variables de décision — État actuel</p>', unsafe_allow_html=True)

    df_vars = pd.DataFrame([
        {"Variable": "x1 — Q_turbine",    "Valeur (T/h)": x1,       "Borne min": 100,  "Borne max": Q_HP_dispo, "Contrainte": "GE : min 100 T/h"},
        {"Variable": "x2 — Bypass HP→MP", "Valeur (T/h)": x2,       "Borne min": 0,    "Borne max": Q_HP_dispo, "Contrainte": "Physique ≥ 0"},
        {"Variable": "x3 — Bypass HP→BP", "Valeur (T/h)": x3,       "Borne min": 0,    "Borne max": Q_HP_dispo, "Contrainte": "Physique ≥ 0"},
        {"Variable": "x4 — Soutirage MP", "Valeur (T/h)": x4,       "Borne min": 0,    "Borne max": 25,         "Contrainte": "Technique 0–25 T/h"},
        {"Variable": "x5 — Soutirage LP", "Valeur (T/h)": x5,       "Borne min": 0,    "Borne max": 130,        "Contrainte": "Technique 0–130 T/h"},
        {"Variable": "x6 — Condenseur",   "Valeur (T/h)": B["x6"],  "Borne min": 0,    "Borne max": "—",        "Contrainte": "x6 = x1−x4−x5 ≥ 0"},
    ])
    st.dataframe(df_vars, use_container_width=True, hide_index=True)

# ══════════════════════════════════════════════
# TAB 2 — ANALYSE DES PERTES
# ══════════════════════════════════════════════
with tab2:
    st.markdown('<p class="section-title">📉 Analyse des Pertes Énergétiques</p>', unsafe_allow_html=True)

    # Pertes calculées
    P_pertes_bypass_MP = x2 * (P["h_HP"] - P["h_MP"]) / 3.6
    P_pertes_bypass_BP = x3 * (P["h_HP"] - P["h_BP"]) / 3.6
    P_pertes_cond      = B["x6"] * (P["h_exhaust"] - 120) / 3.6  # approx
    P_elec_actuel      = B["P_elec_kW"]
    E_totale           = B["E_entree_kW"]

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Bilan énergétique Sankey")
        labels = ["Vapeur HP entrante", "Turbine", "Bypass HP→MP", "Bypass HP→BP",
                  "Puissance électrique", "Condenseur", "Pertes bypass MP", "Pertes bypass BP"]
        fig_sankey = go.Figure(go.Sankey(
            node=dict(
                label=labels,
                color=["#1F4E79","#2196F3","#FF9800","#F44336",
                       "#4CAF50","#9C27B0","#FF9800","#F44336"],
                pad=15, thickness=20
            ),
            link=dict(
                source=[0, 0, 0, 1, 1, 2, 3],
                target=[1, 2, 3, 4, 5, 6, 7],
                value=[max(x1*P["h_HP"]/3.6/1000, 0.1),
                       max(x2*P["h_HP"]/3.6/1000, 0.01),
                       max(x3*P["h_HP"]/3.6/1000, 0.01),
                       max(P_elec_actuel/1000, 0.1),
                       max(P_pertes_cond/1000, 0.1),
                       max(P_pertes_bypass_MP/1000, 0.01),
                       max(P_pertes_bypass_BP/1000, 0.01)],
                color=["rgba(31,78,121,0.3)","rgba(255,152,0,0.3)",
                       "rgba(244,67,54,0.3)","rgba(76,175,80,0.3)",
                       "rgba(156,39,176,0.3)","rgba(255,152,0,0.4)",
                       "rgba(244,67,54,0.4)"]
            )
        ))
        fig_sankey.update_layout(height=350, margin=dict(l=0, r=0, t=20, b=0))
        st.plotly_chart(fig_sankey, use_container_width=True)

    with col2:
        st.subheader("Détail des pertes (kW)")
        df_pertes = pd.DataFrame([
            {"Source de perte":         "Bypass HP→MP (x2)",   "Perte (kW)": round(P_pertes_bypass_MP, 1), "% énergie entrée": round(P_pertes_bypass_MP/max(E_totale,1)*100, 2)},
            {"Source de perte":         "Bypass HP→BP (x3)",   "Perte (kW)": round(P_pertes_bypass_BP, 1), "% énergie entrée": round(P_pertes_bypass_BP/max(E_totale,1)*100, 2)},
            {"Source de perte":         "Condenseur (x6)",     "Perte (kW)": round(P_pertes_cond, 1),      "% énergie entrée": round(P_pertes_cond/max(E_totale,1)*100, 2)},
            {"Source de perte":         "Pertes mécaniques η", "Perte (kW)": round(E_totale*(1-P["eta_gen"])*x1*P["h_HP"]/3.6/max(E_totale,1), 1), "% énergie entrée": round((1-P["eta_gen"])*100, 2)},
        ])
        st.dataframe(df_pertes, use_container_width=True, hide_index=True)

        st.metric("⚠️ Perte totale bypass",
                  f"{(P_pertes_bypass_MP+P_pertes_bypass_BP)/1000:.2f} MW",
                  help="Énergie perdue car vapeur HP détendue sans produire de travail mécanique")

# ══════════════════════════════════════════════
# TAB 3 — DIAGNOSTIC TAGS PI
# ══════════════════════════════════════════════
with tab3:
    st.markdown('<p class="section-title">🔍 Diagnostic Automatique — 43 Tags PI</p>', unsafe_allow_html=True)
    st.info("Saisissez les mesures réelles depuis le DCS Yokogawa pour obtenir le diagnostic automatique.")

    # ── Saisie des mesures par catégories ──
    mesures = {}

    cat_labels = {
        "HP":     "🔴 Admission HP — Collecteur 60 Bar",
        "TRB":    "⚙️ Turbine — GE Nuovo Pignone 3000 RPM",
        "MP":     "🟣 Soutirage MP — Collecteur 12 Bar",
        "BP":     "🩷 Soutirage BP — Collecteur 5 Bar",
        "BP_MP":  "🟠 Détentes Bypass MP",
        "BP_BP":  "🟤 Détentes Bypass BP",
        "COND":   "🔵 Condenseur",
        "VANNE":  "🟢 Vannes — % Ouverture",
        "ARROS":  "💧 Système d'arrosage",
        "CONSO":  "🏭 Consommateurs",
        "ATM":    "🚨 Vannes de sécurité ATM",
    }

    for cat, label in cat_labels.items():
        tags_cat = {k: v for k, v in TAGS.items() if v["cat"] == cat}
        if not tags_cat:
            continue
        with st.expander(f"{label} — {len(tags_cat)} tags", expanded=(cat in ["HP","TRB"])):
            cols = st.columns(3)
            for i, (code, info) in enumerate(tags_cat.items()):
                with cols[i % 3]:
                    tag_display = f'<span class="tag-label">{info["tag"]}</span>'
                    st.markdown(f"{tag_display}", unsafe_allow_html=True)
                    val = st.number_input(
                        info["nom"],
                        value=float(info["nominal"]) if info["nominal"] else 0.0,
                        key=f"mes_{code}",
                        format="%.3f",
                        label_visibility="visible"
                    )
                    mesures[code] = val

    st.divider()

    # ── Résultats diagnostic ──
    if st.button("🔍 Lancer le diagnostic", type="primary", use_container_width=True):
        alertes = diagnostiquer(mesures)

        nb_crit = sum(1 for a in alertes if "CRITIQUE" in a["statut"])
        nb_warn = sum(1 for a in alertes if "WARNING" in a["statut"])
        nb_ok   = sum(1 for a in alertes if "OK" in a["statut"])

        c1, c2, c3 = st.columns(3)
        c1.metric("🔴 CRITIQUE",  nb_crit)
        c2.metric("⚠️ WARNING",   nb_warn)
        c3.metric("✅ OK",         nb_ok)

        if not alertes:
            st.success("Tous les paramètres sont dans les limites nominales ✅")
        else:
            st.subheader("Résultats par ordre de gravité")
            for a in alertes:
                if "CRITIQUE" in a["statut"]:
                    color = "#dc3545"
                elif "WARNING" in a["statut"]:
                    color = "#ffc107"
                else:
                    color = "#28a745"

                with st.expander(
                    f"{a['statut']} — {a['nom']} | Écart : {a['ecart']:+.1f}%",
                    expanded=("CRITIQUE" in a["statut"])
                ):
                    c1, c2, c3 = st.columns(3)
                    c1.metric("Valeur nominale", f"{a['nominal']} {a['unite']}")
                    c2.metric("Valeur mesurée",  f"{a['mesure']} {a['unite']}")
                    c3.metric("Écart",           f"{a['ecart']:+.2f} %")
                    st.markdown(f"**Tag PI :** `{a['tag']}`")
                    st.markdown(f"**Cause probable :** {a['cause']}")
                    st.markdown(f"**Solution recommandée :** {a['sol']}")

# ══════════════════════════════════════════════
# TAB 4 — OPTIMISATION
# ══════════════════════════════════════════════
with tab4:
    st.markdown('<p class="section-title">Optimisation — Maximisation de l\'efficacité énergétique</p>', unsafe_allow_html=True)
    st.markdown("""
    L'optimiseur SLSQP (scipy) cherche les valeurs optimales de **x1, x2, x3, x4, x5**
    qui **maximisent l'efficacité η** tout en respectant toutes les contraintes
    opérationnelles, constructeur GE et de satisfaction des clients.
    """)

    col1, col2 = st.columns([1, 2])

    with col1:
        st.subheader("Paramètres d'optimisation")
        cad_opt   = st.slider("Cadence SAP (%)", 0, 100, int(cadence), key="cad_opt")
        dmp_opt   = st.number_input("Demande MP (T/h)", 0.0, 100.0, float(D_MP), 0.1, key="dmp_opt")
        dbp_opt   = st.number_input("Demande BP (T/h)", 0.0, 200.0, float(D_BP), 1.0, key="dbp_opt")
        hrs_opt   = st.number_input("Apport HRS (T/h)", 0.0, 72.0,  float(Q_HRS), 0.5, key="hrs_opt")

        if st.button(" Optimiser", type="primary", use_container_width=True):
            with st.spinner("Optimisation en cours..."):
                res = optimiser(cad_opt, dmp_opt, dbp_opt, hrs_opt)

            if res["success"]:
                st.success("✅ Optimisation réussie !")
                st.session_state["opt_result"] = res
                st.session_state["opt_params"]  = (cad_opt, dmp_opt, dbp_opt, hrs_opt)
            else:
                st.error("❌ Optimisation échouée — vérifiez les contraintes")

    with col2:
        if "opt_result" in st.session_state:
            res  = st.session_state["opt_result"]
            prms = st.session_state["opt_params"]

            st.subheader("Résultats optimaux")

            c1, c2 = st.columns(2)
            c1.metric(" Puissance optimale",    f"{res['P_elec_MW']:.3f} MW")
            c2.metric(" Efficacité maximale η", f"{res['eta_opt']:.2f} %",
                      f"{res['eta_opt'] - B['eta_glob']:+.2f}% vs simulation actuelle")

            df_opt = pd.DataFrame([
                {"Variable": "x1 — Turbine",    "Actuel (T/h)": x1,    "Optimal (T/h)": res["x1"], "Δ (T/h)": round(res["x1"]-x1, 2)},
                {"Variable": "x2 — Bypass MP",  "Actuel (T/h)": x2,    "Optimal (T/h)": res["x2"], "Δ (T/h)": round(res["x2"]-x2, 2)},
                {"Variable": "x3 — Bypass BP",  "Actuel (T/h)": x3,    "Optimal (T/h)": res["x3"], "Δ (T/h)": round(res["x3"]-x3, 2)},
                {"Variable": "x4 — Soutirage MP","Actuel (T/h)": x4,   "Optimal (T/h)": res["x4"], "Δ (T/h)": round(res["x4"]-x4, 2)},
                {"Variable": "x5 — Soutirage LP","Actuel (T/h)": x5,   "Optimal (T/h)": res["x5"], "Δ (T/h)": round(res["x5"]-x5, 2)},
                {"Variable": "x6 — Condenseur", "Actuel (T/h)": B["x6"],"Optimal (T/h)":res["x6"],"Δ (T/h)": round(res["x6"]-B["x6"], 2)},
            ])

            def color_delta(val):
                if not isinstance(val, float): return ''
                if val > 0.5: return 'background-color:#d4edda'
                if val < -0.5: return 'background-color:#f8d7da'
                return ''

            st.dataframe(
                df_opt.style.map(color_delta, subset=["Δ (T/h)"]),
                use_container_width=True, hide_index=True
            )

            # Graphique comparaison
            fig_comp = go.Figure()
            variables = ["x1", "x2", "x3", "x4", "x5"]
            vals_act  = [x1, x2, x3, x4, x5]
            vals_opt  = [res["x1"], res["x2"], res["x3"], res["x4"], res["x5"]]

            fig_comp.add_trace(go.Bar(name="Actuel", x=variables, y=vals_act, marker_color="#1F4E79"))
            fig_comp.add_trace(go.Bar(name="Optimal", x=variables, y=vals_opt, marker_color="#28a745"))
            fig_comp.update_layout(
                barmode="group", height=280,
                margin=dict(l=0, r=0, t=20, b=0),
                legend=dict(orientation="h", y=1.1)
            )
            st.plotly_chart(fig_comp, use_container_width=True)

        else:
            st.info("Lancez l'optimisation pour voir les résultats.")

    # ── Analyse de sensibilité ──
    st.divider()
    st.subheader("Analyse de sensibilité — Efficacité vs Cadence SAP")

    cadences = list(range(50, 101, 5))
    etas = []
    for c in cadences:
        r = optimiser(c, D_MP, D_BP, Q_HRS)
        etas.append(r["eta_opt"] if r.get("success") else None)

    fig_sens = go.Figure()
    fig_sens.add_trace(go.Scatter(
        x=cadences, y=etas, mode="lines+markers",
        name="η optimal", line=dict(color="#1F4E79", width=2),
        marker=dict(size=6)
    ))
    fig_sens.update_layout(
        xaxis_title="Cadence SAP (%)",
        yaxis_title="Efficacité η (%)",
        height=280,
        margin=dict(l=0, r=0, t=20, b=0)
    )
    st.plotly_chart(fig_sens, use_container_width=True)

# ══════════════════════════════════════════════
# TAB 5 — SCHÉMA P&ID
# ══════════════════════════════════════════════
with tab5:
    st.markdown('<p class="section-title">🗺️ Schéma P&ID — Réseau Vapeur HP/MP/BP</p>', unsafe_allow_html=True)
    st.caption("Cliquez sur un tag PI pour voir ses détails. Schéma DrawIO — Centrale Thermique OCP UNITE 515A")

    file_name = "centrale thermique html.html"

try:
    with open(file_name, "r", encoding="utf-8") as f:
        html_content = f.read()

    st.components.v1.html(html_content, height=950, scrolling=True)

except FileNotFoundError:
    st.error(f"Fichier {file_name} non trouvé")
# ─────────────────────────────────────────────
#  FOOTER
# ─────────────────────────────────────────────
st.divider()
st.markdown("""
<div style="text-align:center;color:#888;font-size:0.8rem;padding:0.5rem">
    EMS — Centrale Thermique OCP | Turbine GE Nuovo Pignone 3000 RPM | UNITE 515A
    <br>Développé dans le cadre du PFE — Master Génie Énergétique
</div>
""", unsafe_allow_html=True)
